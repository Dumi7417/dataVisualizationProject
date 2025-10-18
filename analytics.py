import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule

# 🎨 Seaborn theme
sns.set_theme(style="whitegrid", palette="muted")

# Database connection
conn = psycopg2.connect(
    dbname="spotifydb",
    user="dbuser",
    password="",
    host="localhost",
    port="5432"
)

# Helper to run query
def run_query(query):
    return pd.read_sql(query, conn)

# Visualization Section
def generate_charts():
    print("Generating charts...")

    # 1. Pie chart → Artist distribution by gender (with small slice outside & color marker)
    df = run_query("""
        SELECT gender, COUNT(*) as count
        FROM artists
        GROUP BY gender;
    """)
    total = df['count'].sum()
    df['percentage'] = (df['count'] / total) * 100

    # Split small (<1%) and main slices
    small = df[df['percentage'] < 1]
    df_main = df[df['percentage'] >= 1]

    plt.figure(figsize=(6, 6))
    colors = sns.color_palette("Set2", len(df_main))
    wedges, texts, autotexts = plt.pie(
        df_main['count'],
        labels=df_main['gender'],
        autopct='%1.1f%%',
        colors=colors,
        startangle=90
    )

    # Add small segments outside chart with color box
    if not small.empty:
        x, y = 1.2, 1.0  # starting position for outside labels
        for i, (_, row) in enumerate(small.iterrows()):
            color = sns.color_palette("Set2")[i % len(colors)]
            plt.text(
                x, y - i * 0.1,
                f"{row['gender']} ({row['percentage']:.1f}%)",
                fontsize=10,
                va="center",
            )
            plt.scatter(
                x - 0.08, y - i * 0.1,
                s=120, color=color, marker='s'
            )

    plt.title("Artist Distribution by Gender")
    plt.tight_layout()
    plt.savefig("charts/pie_gender_fixed.png")
    plt.close()
    print("Pie chart saved: charts/pie_gender_fixed.png")

    # 2. Bar chart → Top 10 countries
    df = run_query("""
        SELECT c.country_name, COUNT(*) as artist_count
        FROM countries c
        JOIN artist_countries ac ON c.country_id = ac.country_id
        GROUP BY c.country_name
        ORDER BY artist_count DESC
        LIMIT 10;
    """)
    plt.figure(figsize=(10,6))
    sns.barplot(x="country_name", y="artist_count", data=df, palette="viridis")
    plt.title("Top 10 Countries by Artist Count")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("charts/bar_countries.png")
    plt.close()
    print("Bar chart saved: charts/bar_countries.png")

    # 3. Horizontal bar chart → Top 10 cities
    df = run_query("""
        SELECT ci.city_name, COUNT(*) as artist_count
        FROM cities ci
        JOIN artist_cities ac ON ci.city_id = ac.city_id
        GROUP BY ci.city_name
        ORDER BY artist_count DESC
        LIMIT 10;
    """)
    plt.figure(figsize=(10,6))
    sns.barplot(x="artist_count", y="city_name", data=df, palette="coolwarm")
    plt.title("Top 10 Cities by Artist Count")
    plt.tight_layout()
    plt.savefig("charts/hbar_cities.png")
    plt.close()
    print("Horizontal bar chart saved: charts/hbar_cities.png")

    # 4. Line chart → Average age by gender
    df = run_query("""
        SELECT gender, AVG(age) as avg_age
        FROM artists
        WHERE age IS NOT NULL
        GROUP BY gender
        ORDER BY gender;
    """)
    plt.figure(figsize=(8,5))
    sns.lineplot(x="gender", y="avg_age", data=df, marker="o", color="orange")
    plt.title("Average Age by Gender")
    plt.tight_layout()
    plt.savefig("charts/line_age_gender.png")
    plt.close()
    print("Line chart saved: charts/line_age_gender.png")

    # 5. Histogram → Age distribution
    df = run_query("SELECT age FROM artists WHERE age > 0;")
    plt.figure(figsize=(10,6))
    sns.histplot(df['age'], bins=30, kde=True, color="purple")
    plt.title("Artist Age Distribution")
    plt.tight_layout()
    plt.savefig("charts/hist_age.png")
    plt.close()
    print("Histogram saved: charts/hist_age.png")

    # 6. Scatter plot → Artist Age vs ID (colored by gender)
    df = run_query("SELECT artist_id, age, gender FROM artists WHERE age > 0 LIMIT 500;")
    plt.figure(figsize=(8, 6))
    sns.scatterplot(data=df, x="age", y="artist_id", hue="gender", palette="viridis")
    plt.title("Artist Age vs. ID by Gender")
    plt.xlabel("Artist Age")
    plt.ylabel("Artist (ID proxy)")
    plt.tight_layout()
    plt.savefig("charts/scatter_age_vs_id_fixed.png")
    plt.close()
    print("Scatter plot saved: charts/scatter_age_vs_id_fixed.png")

# Interactive Plotly (Fixed Smooth Slider)
def interactive_plot():
    df = run_query("""
        SELECT a.artist_name, a.age, c.country_name
        FROM artists a
        JOIN artist_countries ac ON a.artist_id = ac.artist_id
        JOIN countries c ON ac.country_id = c.country_id
        WHERE a.age > 0
        LIMIT 300;
    """)
    import numpy as np
    years = list(range(1990, 2023))
    df["year"] = np.random.choice(years, size=len(df))
    all_years = pd.DataFrame({
        "year": years,
        "age": [None]*len(years),
        "country_name": [None]*len(years),
        "artist_name": [None]*len(years)
    })
    df = pd.concat([df, all_years], ignore_index=True)
    df = df.sort_values("year")
    fig = px.scatter(
        df,
        x="age",
        y="country_name",
        animation_frame="year",
        color="country_name",
        hover_name="artist_name",
        size_max=20,
        title="Interactive Age vs Country with Year Slider (1990–2022)"
    )
    fig.show()

# Export to Excel
def export_to_excel():
    df = run_query("SELECT * FROM artists LIMIT 100;")
    wb = Workbook()
    ws = wb.active
    ws.title = "Artists"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(row)
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    rule = ColorScaleRule(start_type="min", start_color="FFAAAA",
                          mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                          end_type="max", end_color="AAFFAA")
    ws.conditional_formatting.add("D2:D100", rule)
    wb.save("exports/spotify_report.xlsx")
    print("Excel export saved: exports/spotify_report.xlsx")

# Main
if __name__ == "__main__":
    generate_charts()
    interactive_plot()
    export_to_excel()
