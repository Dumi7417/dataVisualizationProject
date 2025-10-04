import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# Функция для SQL-запроса
def run_query(conn, query):
    return pd.read_sql(query, conn)

# Экспорт в Excel с форматированием
def export_to_excel(dataframes_dict, filename):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # подсветка для числовых колонок (count, avg_age)
        for col in ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=1):
            if isinstance(col[0].value, str) and (
                "count" in col[0].value.lower() or "avg_age" in col[0].value.lower()
            ):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFCCCC",
                    end_type="max", end_color="00FF00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(filename)
    print(f"✅ File created: {filename} ({len(dataframes_dict)} sheets, rows ~{ws.max_row})")

# --- Основной код ---
conn = psycopg2.connect("dbname=spotifydb user=dbuser")

queries = {
    "ArtistsByGender": """
        SELECT gender, COUNT(*) as count
        FROM artists
        GROUP BY gender;
    """,
    "ArtistsByType": """
        SELECT type, COUNT(*) as count
        FROM artists
        GROUP BY type;
    """,
    "CountriesAll": """
        SELECT c.country_name, COUNT(*) AS artist_count
        FROM countries c
        JOIN artist_countries ac ON c.country_id = ac.country_id
        GROUP BY c.country_name
        ORDER BY artist_count DESC;
    """,
    "CitiesAll": """
        SELECT ci.city_name, c.country_name, COUNT(*) AS artist_count
        FROM cities ci
        JOIN countries c ON ci.country_id = c.country_id
        JOIN artist_cities ac ON ci.city_id = ac.city_id
        GROUP BY ci.city_name, c.country_name
        ORDER BY artist_count DESC;
    """,
    "AvgAgeByGender": """
        SELECT AVG(age) AS avg_age, gender
        FROM artists
        WHERE age IS NOT NULL
        GROUP BY gender;
    """,
    "AvgAgeByCountry": """
        SELECT c.country_name, AVG(age) AS avg_age
        FROM artists a
        JOIN artist_countries ac ON a.artist_id = ac.artist_id
        JOIN countries c ON ac.country_id = c.country_id
        WHERE a.age IS NOT NULL
        GROUP BY c.country_name
        ORDER BY avg_age DESC;
    """
}

# Выполняем запросы и собираем датафреймы
dataframes = {name: run_query(conn, query) for name, query in queries.items()}

# Экспортируем в Excel
export_to_excel(dataframes, "assignment2_results.xlsx")

# Закрываем соединение
conn.close()
